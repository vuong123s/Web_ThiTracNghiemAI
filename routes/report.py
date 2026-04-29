import io
from datetime import datetime
from functools import wraps

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook

from models import Answer, Classroom, Quiz, Submission

report_bp = Blueprint("report", __name__)


def teacher_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_teacher():
            flash("Chức năng báo cáo chỉ dành cho giáo viên.", "warning")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def _build_scoreboard(submissions: list[Submission]) -> list[dict]:
    grouped = {}
    for submission in submissions:
        key = str(submission.student_id) if submission.student_id else submission.display_student_email.lower()
        if key not in grouped:
            grouped[key] = {
                "name": submission.display_student_name,
                "email": submission.display_student_email,
                "scores": [],
                "attempts": 0,
                "latest_at": submission.submitted_at,
            }

        grouped[key]["scores"].append(submission.score_percent)
        grouped[key]["attempts"] += 1
        if submission.submitted_at and (
            grouped[key]["latest_at"] is None or submission.submitted_at > grouped[key]["latest_at"]
        ):
            grouped[key]["latest_at"] = submission.submitted_at

    scoreboard = []
    for item in grouped.values():
        scores = item["scores"]
        scoreboard.append(
            {
                "name": item["name"],
                "email": item["email"],
                "attempts": item["attempts"],
                "avg_score": round(sum(scores) / len(scores), 2) if scores else 0,
                "best_score": round(max(scores), 2) if scores else 0,
                "latest_at": item["latest_at"],
            }
        )

    scoreboard.sort(key=lambda row: (-row["avg_score"], -row["best_score"], row["name"]))
    return scoreboard


def _build_question_stats(selected_quiz: Quiz | None) -> list[dict]:
    if not selected_quiz:
        return []

    stats = []
    for question in selected_quiz.get_ordered_questions():
        base_query = Answer.query.join(Submission).filter(
            Answer.question_id == question.id,
            Submission.quiz_id == selected_quiz.id,
            Submission.status.in_(["submitted", "graded"]),
        )
        total_count = base_query.count()
        correct_count = base_query.filter(Answer.is_correct.is_(True)).count()
        wrong_count = max(total_count - correct_count, 0)
        correct_rate = round((correct_count * 100 / total_count), 1) if total_count else 0

        stats.append(
            {
                "question_id": question.id,
                "order_index": question.order_index,
                "text": question.text,
                "type": question.type,
                "correct_count": correct_count,
                "wrong_count": wrong_count,
                "total_count": total_count,
                "correct_rate": correct_rate,
            }
        )
    return stats


def _get_teacher_report_data(class_id: int | None, quiz_id: int | None):
    classes = (
        Classroom.query.filter_by(teacher_id=current_user.id).order_by(Classroom.created_at.desc()).all()
    )

    quiz_query = Quiz.query.filter_by(teacher_id=current_user.id)
    if class_id:
        quiz_query = quiz_query.filter_by(class_id=class_id)
    quizzes = quiz_query.order_by(Quiz.created_at.desc()).all()
    quiz_ids = [quiz.id for quiz in quizzes]

    selected_quiz = None
    if quiz_id:
        selected_quiz = next((quiz for quiz in quizzes if quiz.id == quiz_id), None)

    if quiz_ids:
        submissions_query = Submission.query.filter(
            Submission.quiz_id.in_(quiz_ids),
            Submission.status.in_(["submitted", "graded"]),
        )
        if selected_quiz:
            submissions_query = submissions_query.filter_by(quiz_id=selected_quiz.id)
        submissions = submissions_query.order_by(Submission.submitted_at.desc()).all()
    else:
        submissions = []

    scoreboard = _build_scoreboard(submissions)
    question_stats = _build_question_stats(selected_quiz)

    return {
        "classes": classes,
        "quizzes": quizzes,
        "selected_quiz": selected_quiz,
        "submissions": submissions,
        "scoreboard": scoreboard,
        "question_stats": question_stats,
    }


@report_bp.route("/")
@login_required
def dashboard():
    if current_user.is_student():
        return redirect(url_for("quiz.history"))

    if not current_user.is_teacher():
        flash("Không có dữ liệu báo cáo cho tài khoản này.", "warning")
        return redirect(url_for("dashboard"))

    class_id = request.args.get("class_id", type=int)
    quiz_id = request.args.get("quiz_id", type=int)
    data = _get_teacher_report_data(class_id=class_id, quiz_id=quiz_id)

    chart_labels = [
        f"Câu {idx + 1}" for idx, _ in enumerate(data["question_stats"])
    ]
    chart_correct = [row["correct_count"] for row in data["question_stats"]]
    chart_wrong = [row["wrong_count"] for row in data["question_stats"]]

    return render_template(
        "report/dashboard.html",
        classes=data["classes"],
        quizzes=data["quizzes"],
        selected_class_id=class_id,
        selected_quiz_id=quiz_id,
        selected_quiz=data["selected_quiz"],
        submissions=data["submissions"][:30],
        scoreboard=data["scoreboard"],
        question_stats=data["question_stats"],
        chart_labels=chart_labels,
        chart_correct=chart_correct,
        chart_wrong=chart_wrong,
    )


@report_bp.route("/export")
@teacher_required
def export_report():
    class_id = request.args.get("class_id", type=int)
    quiz_id = request.args.get("quiz_id", type=int)
    data = _get_teacher_report_data(class_id=class_id, quiz_id=quiz_id)

    workbook = Workbook()
    score_sheet = workbook.active
    score_sheet.title = "Scoreboard"
    score_sheet.append(["Student Name", "Email", "Attempts", "Average Score (%)", "Best Score (%)"])
    for row in data["scoreboard"]:
        score_sheet.append(
            [row["name"], row["email"], row["attempts"], row["avg_score"], row["best_score"]]
        )

    submission_sheet = workbook.create_sheet("Submissions")
    submission_sheet.append(["Quiz", "Student", "Email", "Score", "Max", "Percent", "Submitted At"])
    for row in data["submissions"]:
        submission_sheet.append(
            [
                row.quiz.title,
                row.display_student_name,
                row.display_student_email,
                row.total_score,
                row.max_score,
                row.score_percent,
                row.submitted_at.strftime("%Y-%m-%d %H:%M") if row.submitted_at else "",
            ]
        )

    if data["selected_quiz"]:
        question_sheet = workbook.create_sheet("Question Stats")
        question_sheet.append(["Order", "Question", "Correct", "Wrong", "Total", "Correct Rate (%)"])
        for row in data["question_stats"]:
            question_sheet.append(
                [
                    row["order_index"],
                    row["text"],
                    row["correct_count"],
                    row["wrong_count"],
                    row["total_count"],
                    row["correct_rate"],
                ]
            )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    selected_name = data["selected_quiz"].title if data["selected_quiz"] else "all"
    safe_name = "".join(ch if ch.isalnum() else "_" for ch in selected_name).strip("_") or "all"
    filename = f"teacher_report_{safe_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
