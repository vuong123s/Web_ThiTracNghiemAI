import io
import random
import re
from datetime import datetime
from functools import wraps
from urllib.parse import quote

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook

from forms import EssayGradingForm, ImportBankForm, QuestionBankForm, QuizBuilderForm, UserForm
from models import (
    Answer,
    BankOption,
    Classroom,
    Question,
    QuestionBank,
    Quiz,
    Submission,
    User,
    db,
)

admin_bp = Blueprint("admin", __name__)


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin():
            flash("Bạn không có quyền truy cập trang này.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def _parse_datetime(raw_value: str | None):
    if not raw_value:
        return None
    value = raw_value.strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _normalize_difficulty(value: str | None, default: str = "medium") -> str:
    if not value:
        return default
    normalized = value.strip().lower()
    if normalized in {"easy", "medium", "hard"}:
        return normalized
    return default


def _extract_correct_letters(raw_value: str | None) -> set[str]:
    if not raw_value:
        return set()
    letters = set()
    for chunk in raw_value.split(","):
        value = chunk.strip().upper()
        if value in {"A", "B", "C", "D"}:
            letters.add(value)
    return letters


def _sync_bank_options(bank_question: QuestionBank, form: QuestionBankForm) -> None:
    bank_question.options.delete()

    if bank_question.type == "essay":
        return

    option_payloads = [
        ("A", form.option_a.data),
        ("B", form.option_b.data),
        ("C", form.option_c.data),
        ("D", form.option_d.data),
    ]
    correct_letters = _extract_correct_letters(form.correct_answers.data)

    for idx, (letter, text) in enumerate(option_payloads, start=1):
        if not text or not text.strip():
            continue
        db.session.add(
            BankOption(
                question_bank_id=bank_question.id,
                text=text.strip(),
                is_correct=letter in correct_letters,
                order_index=idx,
            )
        )


def _import_from_excel(file_storage, teacher_id: int, default_category: str, default_difficulty: str):
    workbook = load_workbook(file_storage)
    sheet = workbook.active

    imported = 0
    errors = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        text = str(row[0]).strip()
        qtype = (str(row[1]).strip().lower() if len(row) > 1 and row[1] else "single")
        category = (str(row[2]).strip() if len(row) > 2 and row[2] else default_category) or None
        difficulty = _normalize_difficulty(
            str(row[3]).strip() if len(row) > 3 and row[3] else default_difficulty
        )
        tags = str(row[4]).strip() if len(row) > 4 and row[4] else None
        option_a = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        option_b = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        option_c = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        option_d = str(row[8]).strip() if len(row) > 8 and row[8] else ""
        correct_raw = str(row[9]).strip() if len(row) > 9 and row[9] else ""
        explanation = str(row[10]).strip() if len(row) > 10 and row[10] else None

        if qtype not in {"single", "multiple", "essay"}:
            errors.append(f"Dòng {row_index}: type không hợp lệ ({qtype})")
            continue

        item = QuestionBank(
            teacher_id=teacher_id,
            text=text,
            type=qtype,
            category=category,
            difficulty=difficulty,
            tags=tags,
            explanation=explanation,
        )
        db.session.add(item)
        db.session.flush()

        if qtype != "essay":
            options = [("A", option_a), ("B", option_b), ("C", option_c), ("D", option_d)]
            correct_letters = _extract_correct_letters(correct_raw)
            non_empty_count = 0
            for order, (letter, opt_text) in enumerate(options, start=1):
                if not opt_text:
                    continue
                non_empty_count += 1
                db.session.add(
                    BankOption(
                        question_bank_id=item.id,
                        text=opt_text,
                        is_correct=letter in correct_letters,
                        order_index=order,
                    )
                )
            if non_empty_count < 2:
                errors.append(f"Dòng {row_index}: câu trắc nghiệm cần ít nhất 2 đáp án.")
                db.session.delete(item)
                continue

        imported += 1

    return imported, errors


def _import_from_docx(file_storage, teacher_id: int, default_category: str, default_difficulty: str):
    try:
        from docx import Document
    except ImportError as exc:
        raise RuntimeError("Thiếu thư viện python-docx. Hãy cài: pip install python-docx") from exc

    document = Document(file_storage)
    lines = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]

    imported = 0
    errors = []
    current = None

    def flush_current_question():
        nonlocal imported, current
        if not current:
            return

        question_text = current.get("text", "").strip()
        if not question_text:
            current = None
            return

        qtype = current.get("type", "single")
        item = QuestionBank(
            teacher_id=teacher_id,
            text=question_text,
            type=qtype,
            category=default_category or None,
            difficulty=default_difficulty,
            tags=current.get("tags"),
            explanation=current.get("explanation"),
        )
        db.session.add(item)
        db.session.flush()

        if qtype != "essay":
            options = current.get("options", {})
            correct_letters = _extract_correct_letters(current.get("correct"))
            for idx, letter in enumerate(["A", "B", "C", "D"], start=1):
                if letter in options:
                    db.session.add(
                        BankOption(
                            question_bank_id=item.id,
                            text=options[letter],
                            is_correct=letter in correct_letters,
                            order_index=idx,
                        )
                    )

        imported += 1
        current = None

    for line in lines:
        if line.upper().startswith("Q:"):
            flush_current_question()
            current = {"text": line[2:].strip(), "type": "single", "options": {}}
            continue

        if line.upper().startswith("TYPE:") and current:
            qtype = line.split(":", 1)[1].strip().lower()
            if qtype in {"single", "multiple", "essay"}:
                current["type"] = qtype
            continue

        if line.upper().startswith("ANS:") and current:
            current["correct"] = line.split(":", 1)[1].strip()
            continue

        if line.upper().startswith("EXPL:") and current:
            current["explanation"] = line.split(":", 1)[1].strip()
            continue

        option_match = re.match(r"^([A-Da-d])[\.\):\-]\s*(.+)$", line)
        if option_match and current:
            letter = option_match.group(1).upper()
            current["options"][letter] = option_match.group(2).strip()
            continue

        if current:
            current["text"] = f"{current['text']} {line}".strip()
        else:
            item = QuestionBank(
                teacher_id=teacher_id,
                text=line,
                type="essay",
                category=default_category or None,
                difficulty=default_difficulty,
            )
            db.session.add(item)
            imported += 1

    flush_current_question()
    return imported, errors


# ==================== Dashboard ====================
@admin_bp.route("/")
@admin_required
def dashboard():
    total_questions = QuestionBank.query.count()
    total_quizzes = Quiz.query.count()
    total_students = User.query.filter_by(role="student").count()
    total_teachers = User.query.filter_by(role="teacher").count()
    total_results = Submission.query.filter(Submission.status.in_(["submitted", "graded"])).count()

    recent_quizzes = Quiz.query.order_by(Quiz.created_at.desc()).limit(5).all()
    recent_results = (
        Submission.query.filter(Submission.status.in_(["submitted", "graded"]))
        .order_by(Submission.submitted_at.desc())
        .limit(10)
        .all()
    )

    scores = [row.score_percent for row in recent_results if row.max_score > 0]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0

    return render_template(
        "admin/dashboard.html",
        total_questions=total_questions,
        total_quizzes=total_quizzes,
        total_students=total_students,
        total_teachers=total_teachers,
        total_results=total_results,
        recent_quizzes=recent_quizzes,
        recent_results=recent_results,
        avg_score=avg_score,
    )


# ==================== Question Bank ====================
@admin_bp.route("/questions")
@admin_required
def questions_list():
    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()
    difficulty = request.args.get("difficulty", "").strip()
    teacher_id = request.args.get("teacher_id", type=int)

    query = QuestionBank.query
    if search:
        query = query.filter(QuestionBank.text.ilike(f"%{search}%"))
    if category:
        query = query.filter(QuestionBank.category.ilike(f"%{category}%"))
    if difficulty in {"easy", "medium", "hard"}:
        query = query.filter_by(difficulty=difficulty)
    if teacher_id:
        query = query.filter_by(teacher_id=teacher_id)

    questions = query.order_by(QuestionBank.created_at.desc()).all()
    categories = (
        db.session.query(QuestionBank.category)
        .filter(QuestionBank.category.isnot(None))
        .distinct()
        .all()
    )
    categories = [c[0] for c in categories if c[0]]
    teachers = User.query.filter(User.role.in_(["teacher", "admin"])).order_by(User.full_name.asc()).all()

    return render_template(
        "admin/questions/list.html",
        questions=questions,
        categories=categories,
        teachers=teachers,
        current_search=search,
        current_category=category,
        current_difficulty=difficulty,
        current_teacher_id=teacher_id,
    )


@admin_bp.route("/questions/create", methods=["GET", "POST"])
@admin_required
def questions_create():
    form = QuestionBankForm()
    if form.validate_on_submit():
        question = QuestionBank(
            teacher_id=current_user.id,
            text=form.text.data,
            type=form.type.data,
            category=form.category.data or None,
            difficulty=form.difficulty.data,
            tags=form.tags.data or None,
            explanation=form.explanation.data or None,
        )
        db.session.add(question)
        db.session.flush()
        _sync_bank_options(question, form)
        db.session.commit()
        flash("Đã tạo câu hỏi trong ngân hàng.", "success")
        return redirect(url_for("admin.questions_list"))

    return render_template("admin/questions/create.html", form=form)


@admin_bp.route("/questions/<int:id>/edit", methods=["GET", "POST"])
@admin_required
def questions_edit(id):
    question = QuestionBank.query.get_or_404(id)
    form = QuestionBankForm(obj=question)
    existing = {opt.order_index: opt for opt in question.options.all()}

    if request.method == "GET":
        form.option_a.data = existing.get(1).text if existing.get(1) else ""
        form.option_b.data = existing.get(2).text if existing.get(2) else ""
        form.option_c.data = existing.get(3).text if existing.get(3) else ""
        form.option_d.data = existing.get(4).text if existing.get(4) else ""
        form.correct_answers.data = ",".join(
            [opt.label for opt in question.options.filter_by(is_correct=True)]
        )

    if form.validate_on_submit():
        question.text = form.text.data
        question.type = form.type.data
        question.category = form.category.data or None
        question.difficulty = form.difficulty.data
        question.tags = form.tags.data or None
        question.explanation = form.explanation.data or None
        _sync_bank_options(question, form)
        db.session.commit()
        flash("Cập nhật câu hỏi thành công.", "success")
        return redirect(url_for("admin.questions_list"))

    return render_template("admin/questions/edit.html", form=form, question=question)


@admin_bp.route("/questions/<int:id>/delete", methods=["POST"])
@admin_required
def questions_delete(id):
    question = QuestionBank.query.get_or_404(id)
    db.session.delete(question)
    db.session.commit()
    flash("Đã xóa câu hỏi khỏi ngân hàng.", "success")
    return redirect(url_for("admin.questions_list"))


@admin_bp.route("/questions/import", methods=["GET", "POST"])
@admin_required
def questions_import():
    form = ImportBankForm()
    if form.validate_on_submit():
        file = form.file.data
        filename = (file.filename or "").lower()
        default_category = form.default_category.data.strip() if form.default_category.data else ""
        default_difficulty = _normalize_difficulty(form.default_difficulty.data)

        try:
            if filename.endswith((".xlsx", ".xls")):
                imported, errors = _import_from_excel(
                    file, current_user.id, default_category, default_difficulty
                )
            elif filename.endswith(".docx"):
                imported, errors = _import_from_docx(
                    file, current_user.id, default_category, default_difficulty
                )
            else:
                flash("File không hợp lệ. Chỉ hỗ trợ .xlsx/.xls/.docx", "danger")
                return render_template("admin/questions/import.html", form=form)

            db.session.commit()
            if imported:
                flash(f"Import thành công {imported} câu hỏi.", "success")
            if errors:
                flash("Một số dòng bị bỏ qua: " + "; ".join(errors[:5]), "warning")
            return redirect(url_for("admin.questions_list"))
        except Exception as exc:
            db.session.rollback()
            flash(f"Lỗi import: {exc}", "danger")

    return render_template("admin/questions/import.html", form=form)


@admin_bp.route("/questions/export")
@admin_required
def questions_export():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "QuestionBank"
    sheet.append(
        [
            "text",
            "type",
            "category",
            "difficulty",
            "tags",
            "option_a",
            "option_b",
            "option_c",
            "option_d",
            "correct_answers",
            "explanation",
        ]
    )

    rows = QuestionBank.query.order_by(QuestionBank.id.asc()).all()
    for item in rows:
        options = {opt.order_index: opt for opt in item.options.all()}
        correct = ",".join([opt.label for opt in item.options.filter_by(is_correct=True).all()])
        sheet.append(
            [
                item.text,
                item.type,
                item.category or "",
                item.difficulty,
                item.tags or "",
                options.get(1).text if options.get(1) else "",
                options.get(2).text if options.get(2) else "",
                options.get(3).text if options.get(3) else "",
                options.get(4).text if options.get(4) else "",
                correct,
                item.explanation or "",
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="question_bank_all.xlsx",
    )


@admin_bp.route("/questions/template")
@admin_required
def questions_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Template"
    sheet.append(
        [
            "text",
            "type",
            "category",
            "difficulty",
            "tags",
            "option_a",
            "option_b",
            "option_c",
            "option_d",
            "correct_answers",
            "explanation",
        ]
    )
    sheet.append(
        [
            "Python là ngôn ngữ gì?",
            "single",
            "Tin học",
            "easy",
            "python,co-ban",
            "Ngôn ngữ biên dịch",
            "Ngôn ngữ thông dịch",
            "Ngôn ngữ máy",
            "Ngôn ngữ đánh dấu",
            "B",
            "Python là ngôn ngữ thông dịch bậc cao.",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="question_bank_template.xlsx",
    )


# ==================== Quiz Management ====================
@admin_bp.route("/quizzes")
@admin_required
def quizzes_list():
    quizzes = Quiz.query.order_by(Quiz.created_at.desc()).all()
    return render_template("admin/quizzes/list.html", quizzes=quizzes)


@admin_bp.route("/quizzes/create", methods=["GET", "POST"])
@admin_required
def quizzes_create():
    form = QuizBuilderForm()
    class_choices = [(0, "Không gán lớp")] + [
        (classroom.id, f"{classroom.name} ({classroom.join_code})")
        for classroom in Classroom.query.order_by(Classroom.name.asc()).all()
    ]
    form.class_id.choices = class_choices

    available_bank_questions = QuestionBank.query.order_by(QuestionBank.id.desc()).all()

    if form.validate_on_submit():
        scheduled_at = _parse_datetime(form.scheduled_at.data)
        expires_at = _parse_datetime(form.expires_at.data)

        class_id = form.class_id.data if form.class_id.data else None
        teacher_id = current_user.id
        if class_id:
            classroom = Classroom.query.get(class_id)
            if classroom:
                teacher_id = classroom.teacher_id

        quiz = Quiz(
            title=form.title.data,
            description=form.description.data or None,
            time_limit_minutes=form.time_limit_minutes.data,
            is_shuffled=form.is_shuffled.data,
            shuffle_options=form.shuffle_options.data,
            is_published=form.is_published.data,
            allow_retake=form.allow_retake.data,
            teacher_id=teacher_id,
            class_id=class_id,
            scheduled_at=scheduled_at,
            expires_at=expires_at,
            brand_logo_url=form.brand_logo_url.data or None,
            brand_color=form.brand_color.data or None,
        )
        db.session.add(quiz)
        db.session.flush()

        selected_bank_questions = []
        if form.source_mode.data == "random":
            random_query = QuestionBank.query
            if form.random_category.data:
                random_query = random_query.filter(
                    QuestionBank.category.ilike(f"%{form.random_category.data.strip()}%")
                )
            if form.random_difficulty.data:
                random_query = random_query.filter_by(difficulty=form.random_difficulty.data)
            candidates = random_query.all()
            requested_count = form.random_count.data or 10
            selected_bank_questions = (
                random.sample(candidates, requested_count)
                if len(candidates) >= requested_count
                else candidates
            )
        else:
            selected_ids = [
                int(raw_id)
                for raw_id in request.form.getlist("bank_question_ids")
                if raw_id.isdigit()
            ]
            if selected_ids:
                selected_bank_questions = (
                    QuestionBank.query.filter(QuestionBank.id.in_(selected_ids))
                    .order_by(QuestionBank.id.asc())
                    .all()
                )

        for idx, bank_item in enumerate(selected_bank_questions, start=1):
            db.session.add(bank_item.clone_to_quiz(quiz.id, order_index=idx, created_by=current_user.id))

        db.session.commit()
        flash(f"Tạo đề thi thành công với {len(selected_bank_questions)} câu.", "success")
        return redirect(url_for("admin.quizzes_list"))

    return render_template(
        "admin/quizzes/create.html",
        form=form,
        bank_questions=available_bank_questions,
    )


@admin_bp.route("/quizzes/<int:id>/edit", methods=["GET", "POST"])
@admin_required
def quizzes_edit(id):
    quiz = Quiz.query.get_or_404(id)
    form = QuizBuilderForm(obj=quiz)
    class_choices = [(0, "Không gán lớp")] + [
        (classroom.id, f"{classroom.name} ({classroom.join_code})")
        for classroom in Classroom.query.order_by(Classroom.name.asc()).all()
    ]
    form.class_id.choices = class_choices

    if request.method == "GET":
        form.class_id.data = quiz.class_id or 0
        form.scheduled_at.data = quiz.scheduled_at.strftime("%Y-%m-%d %H:%M") if quiz.scheduled_at else ""
        form.expires_at.data = quiz.expires_at.strftime("%Y-%m-%d %H:%M") if quiz.expires_at else ""

    if form.validate_on_submit():
        quiz.title = form.title.data
        quiz.description = form.description.data or None
        quiz.time_limit_minutes = form.time_limit_minutes.data
        quiz.is_shuffled = form.is_shuffled.data
        quiz.shuffle_options = form.shuffle_options.data
        quiz.is_published = form.is_published.data
        quiz.allow_retake = form.allow_retake.data

        new_class_id = form.class_id.data if form.class_id.data else None
        quiz.class_id = new_class_id
        if new_class_id:
            classroom = Classroom.query.get(new_class_id)
            if classroom:
                quiz.teacher_id = classroom.teacher_id

        quiz.scheduled_at = _parse_datetime(form.scheduled_at.data)
        quiz.expires_at = _parse_datetime(form.expires_at.data)
        quiz.brand_logo_url = form.brand_logo_url.data or None
        quiz.brand_color = form.brand_color.data or None

        add_ids = [int(raw_id) for raw_id in request.form.getlist("bank_question_ids") if raw_id.isdigit()]
        if add_ids:
            existing_count = quiz.questions.count()
            additions = (
                QuestionBank.query.filter(QuestionBank.id.in_(add_ids))
                .order_by(QuestionBank.id.asc())
                .all()
            )
            for offset, bank_item in enumerate(additions, start=1):
                db.session.add(
                    bank_item.clone_to_quiz(
                        quiz.id,
                        order_index=existing_count + offset,
                        created_by=current_user.id,
                    )
                )

        db.session.commit()
        flash("Cập nhật đề thi thành công.", "success")
        return redirect(url_for("admin.quizzes_list"))

    bank_questions = QuestionBank.query.order_by(QuestionBank.id.desc()).all()
    quiz_questions = quiz.get_ordered_questions()
    return render_template(
        "admin/quizzes/edit.html",
        form=form,
        quiz=quiz,
        bank_questions=bank_questions,
        quiz_questions=quiz_questions,
    )


@admin_bp.route("/quizzes/<int:id>/questions/<int:question_id>/media", methods=["POST"])
@admin_required
def quiz_question_media_update(id, question_id):
    quiz = Quiz.query.get_or_404(id)
    question = Question.query.get_or_404(question_id)
    if question.quiz_id != quiz.id:
        flash("Câu hỏi không thuộc đề thi này.", "danger")
        return redirect(url_for("admin.quizzes_edit", id=id))

    image_url = (request.form.get("image_url") or "").strip() or None
    youtube_url = (request.form.get("youtube_url") or "").strip() or None

    if youtube_url and "youtube.com" not in youtube_url and "youtu.be" not in youtube_url:
        flash("Link YouTube không hợp lệ.", "danger")
        return redirect(url_for("admin.quizzes_edit", id=id))

    question.image_url = image_url
    question.youtube_url = youtube_url
    db.session.commit()
    flash("Đã cập nhật media cho câu hỏi.", "success")
    return redirect(url_for("admin.quizzes_edit", id=id))


@admin_bp.route("/quizzes/<int:id>/delete", methods=["POST"])
@admin_required
def quizzes_delete(id):
    quiz = Quiz.query.get_or_404(id)
    db.session.delete(quiz)
    db.session.commit()
    flash("Đã xóa đề thi.", "success")
    return redirect(url_for("admin.quizzes_list"))


@admin_bp.route("/quizzes/<int:id>/regenerate", methods=["POST"])
@admin_required
def quizzes_regenerate(id):
    quiz = Quiz.query.get_or_404(id)

    random_count = request.form.get("random_count", type=int) or max(quiz.questions.count(), 5)
    candidates = QuestionBank.query.all()
    selected = random.sample(candidates, random_count) if len(candidates) >= random_count else candidates

    quiz.questions.delete()
    for idx, bank_item in enumerate(selected, start=1):
        db.session.add(bank_item.clone_to_quiz(quiz.id, order_index=idx, created_by=current_user.id))
    db.session.commit()

    flash(f"Đã tạo lại đề thi với {len(selected)} câu hỏi.", "success")
    return redirect(url_for("admin.quizzes_edit", id=id))


@admin_bp.route("/quizzes/<int:id>/export")
@admin_required
def quizzes_export(id):
    quiz = Quiz.query.get_or_404(id)

    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "Quiz"
    info_sheet.append(["Title", quiz.title])
    info_sheet.append(["Description", quiz.description or ""])
    info_sheet.append(["Class", quiz.classroom.name if quiz.classroom else ""])
    info_sheet.append(["TimeLimit", quiz.time_limit_minutes])
    info_sheet.append(["ShuffleQuestions", "Yes" if quiz.is_shuffled else "No"])
    info_sheet.append(["ShuffleOptions", "Yes" if quiz.shuffle_options else "No"])
    info_sheet.append(["Published", "Yes" if quiz.is_published else "No"])

    question_sheet = workbook.create_sheet("Questions")
    question_sheet.append(["Order", "Type", "Question", "CorrectOptions", "Options", "Explanation"])
    for question in quiz.get_ordered_questions():
        options = question.options.order_by("order_index").all()
        options_text = " | ".join([f"{opt.label}. {opt.text}" for opt in options])
        correct_letters = ",".join([opt.label for opt in question.correct_options])
        question_sheet.append(
            [
                question.order_index,
                question.type,
                question.text,
                correct_letters,
                options_text,
                question.explanation or "",
            ]
        )

    result_sheet = workbook.create_sheet("Submissions")
    result_sheet.append(["Student", "Email", "Score", "MaxScore", "Percent", "Status", "SubmittedAt"])
    for sub in quiz.submissions.order_by(Submission.submitted_at.desc()).all():
        result_sheet.append(
            [
                sub.display_student_name,
                sub.display_student_email,
                sub.total_score,
                sub.max_score,
                sub.score_percent,
                sub.status,
                sub.submitted_at.strftime("%d/%m/%Y %H:%M") if sub.submitted_at else "",
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"quiz_{quiz.id}.xlsx",
    )


@admin_bp.route("/quizzes/<int:id>/share")
@admin_required
def quizzes_share(id):
    quiz = Quiz.query.get_or_404(id)
    token = quiz.ensure_share_token()
    db.session.commit()

    share_url = url_for("quiz.shared_quiz_entry", token=token, _external=True)
    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=260x260&data={quote(share_url)}"
    return render_template("teacher/quizzes/share.html", quiz=quiz, share_url=share_url, qr_url=qr_url)


# ==================== User Management ====================
@admin_bp.route("/users")
@admin_required
def users_list():
    search = request.args.get("search", "").strip()
    role = request.args.get("role", "").strip()

    query = User.query
    if search:
        query = query.filter(
            (User.full_name.ilike(f"%{search}%")) | (User.email.ilike(f"%{search}%"))
        )
    if role:
        query = query.filter_by(role=role)

    users = query.order_by(User.created_at.desc()).all()
    return render_template(
        "admin/users/list.html",
        users=users,
        current_search=search,
        current_role=role,
    )


@admin_bp.route("/users/create", methods=["GET", "POST"])
@admin_required
def users_create():
    form = UserForm()

    if form.validate_on_submit():
        existing = User.query.filter_by(email=form.email.data).first()
        if existing:
            flash("Email này đã được sử dụng.", "danger")
            return render_template("admin/users/create.html", form=form)

        user = User(
            full_name=form.full_name.data,
            email=form.email.data,
            role=form.role.data,
        )
        user.set_password(form.password.data or "123456")
        db.session.add(user)
        db.session.commit()
        flash("Tạo tài khoản thành công!", "success")
        return redirect(url_for("admin.users_list"))

    return render_template("admin/users/create.html", form=form)


@admin_bp.route("/users/<int:id>/edit", methods=["GET", "POST"])
@admin_required
def users_edit(id):
    user = User.query.get_or_404(id)
    form = UserForm(obj=user)

    if form.validate_on_submit():
        if form.email.data != user.email:
            existing = User.query.filter_by(email=form.email.data).first()
            if existing:
                flash("Email này đã được sử dụng.", "danger")
                return render_template("admin/users/edit.html", form=form, user=user)

        user.full_name = form.full_name.data
        user.email = form.email.data
        user.role = form.role.data

        if form.password.data:
            user.set_password(form.password.data)

        db.session.commit()
        flash("Cập nhật tài khoản thành công!", "success")
        return redirect(url_for("admin.users_list"))

    return render_template("admin/users/edit.html", form=form, user=user)


@admin_bp.route("/users/<int:id>/delete", methods=["POST"])
@admin_required
def users_delete(id):
    user = User.query.get_or_404(id)

    if user.id == current_user.id:
        flash("Bạn không thể xóa tài khoản của chính mình.", "danger")
        return redirect(url_for("admin.users_list"))

    db.session.delete(user)
    db.session.commit()
    flash("Xóa tài khoản thành công!", "success")
    return redirect(url_for("admin.users_list"))


@admin_bp.route("/users/<int:id>/toggle-role", methods=["POST"])
@admin_required
def users_toggle_role(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash("Bạn không thể thay đổi vai trò của chính mình.", "warning")
    else:
        new_role = request.form.get("role")
        if new_role in ["admin", "teacher", "student"]:
            user.role = new_role
            db.session.commit()
            flash(f"Đã thay đổi vai trò của {user.full_name} thành {new_role}.", "success")
    return redirect(url_for("admin.users_list"))


# ==================== Results/Statistics ====================
@admin_bp.route("/results")
@admin_required
def results():
    quiz_id = request.args.get("quiz_id", type=int)
    query = Submission.query.filter(Submission.status.in_(["submitted", "graded"]))
    if quiz_id:
        query = query.filter_by(quiz_id=quiz_id)

    results = query.order_by(Submission.submitted_at.desc()).all()
    quizzes = Quiz.query.order_by(Quiz.created_at.desc()).all()

    return render_template(
        "admin/results/list.html",
        results=results,
        quizzes=quizzes,
        current_quiz_id=quiz_id,
    )


@admin_bp.route("/results/<int:result_id>")
@admin_required
def result_detail(result_id):
    submission = Submission.query.get_or_404(result_id)
    grading_forms = {answer.id: EssayGradingForm() for answer in submission.answers.all()}
    return render_template(
        "admin/results/detail.html",
        result=submission,
        user_answers=submission.answers.order_by(Answer.id.asc()).all(),
        grading_forms=grading_forms,
    )


@admin_bp.route("/results/<int:result_id>/grade/<int:answer_id>", methods=["POST"])
@admin_required
def grade_essay(result_id, answer_id):
    submission = Submission.query.get_or_404(result_id)
    answer = Answer.query.get_or_404(answer_id)
    if answer.submission_id != submission.id or answer.question.type != "essay":
        flash("Không thể chấm câu trả lời này.", "danger")
        return redirect(url_for("admin.result_detail", result_id=result_id))

    form = EssayGradingForm()
    if form.validate_on_submit():
        answer.apply_manual_grading(score=form.score_awarded.data, feedback=form.feedback.data)
        answer.is_correct = answer.score_awarded >= max(answer.question.points, 1)

        submission.status = "graded"
        submission.recalculate_scores()
        if not submission.submitted_at:
            submission.submitted_at = datetime.utcnow()
        db.session.commit()
        flash("Đã lưu chấm điểm tự luận.", "success")
    else:
        flash("Thông tin chấm điểm chưa hợp lệ.", "danger")

    return redirect(url_for("admin.result_detail", result_id=result_id))
